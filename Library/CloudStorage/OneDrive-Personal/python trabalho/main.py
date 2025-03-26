import openpyxl

# Carregar o arquivo Excel existente
input_file = "itens_python.xlsx"  # Substitua pelo nome do seu arquivo
wb = openpyxl.load_workbook(input_file)
sheet = wb.active

# Criar um novo arquivo Excel para os dados processados
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active
new_sheet.title = "Dados Processados"

# Criar cabeçalhos na nova planilha
new_sheet.append(["Itens", "Descrição Detalhada", "Quantidade Total", "Valor Unitário", "Unidade de Fornecimento"])

# Iterar pelas linhas ímpares e pares da planilha original
for i in range(1, sheet.max_row + 1, 2):  # Processar apenas linhas ímpares
    item_number = sheet.cell(row=i, column=1).value  # Número do item
    descricao_detalhada = None  # Inicializar como None
    quantidade_total = valor_unitario = unidade_fornecimento = None  # Inicializar valores padrão

    # Verificar a linha par seguinte para informações adicionais
    if i + 1 <= sheet.max_row:  # Garantir que a linha par existe
        # Obter os dados originais da célula na linha par
        original_cell_value = sheet.cell(row=i + 1, column=1).value
        if isinstance(original_cell_value, str):  # Garantir que é uma string
            # Processar "Descrição Detalhada"
            if "Descrição Detalhada:" in original_cell_value:
                descricao_detalhada = original_cell_value.split("Descrição Detalhada:")[1].split("Tratamento Diferenciado:")[0].strip()
            # Processar "Quantidade Total"
            if "Quantidade Total:" in original_cell_value:
                quantidade_total = original_cell_value.split("Quantidade Total:")[1].split("Quantidade Mínima Cotada:")[0].strip()
            # Processar "Valor Unitário"

            if "Valor Unitário (R$):" in original_cell_value:
                valor_unitario = original_cell_value.split("Valor Unitário (R$):")[1].split("Unidade de Fornecimento:")[0].strip()
            # Processar "Unidade de Fornecimento"
            if "Unidade de Fornecimento:" in original_cell_value:
                unidade_fornecimento = original_cell_value.split("Unidade de Fornecimento:")[1].split("Quantidade Máxima para Adesões:")[0].strip()

    # Adicionar os dados processados na nova planilha
    new_sheet.append([item_number, descricao_detalhada, quantidade_total, valor_unitario, unidade_fornecimento])

# Salvar o novo arquivo Excel
output_file = "dados_processados.xlsx"
new_wb.save(output_file)
print(f"Arquivo processado salvo como: {output_file}")

